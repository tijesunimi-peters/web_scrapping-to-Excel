from openpyxl import load_workbook
from bs4 import BeautifulSoup
from re import escape
from requests import get, codes
import logging
import argparse
import time

log = logging.getLogger() # 'root' Logger
console = logging.StreamHandler()
format_str = '%(asctime)s\t%(levelname)s -- %(filename)s:%(lineno)s -- %(message)s'
console.setFormatter(logging.Formatter(format_str))
log.addHandler(console) # prints to console.
log.setLevel(logging.DEBUG) # anything ERROR or above


def get_args():
    '''This function parses and return arguments passed in'''
    parser = argparse.ArgumentParser(description='Script retrieves schedules from a given server')
    parser.add_argument('-s', '--state', type=str, help='State to find job. leave blank for all states', default=None)
    parser.add_argument('-se', '--sector', type=str, help='Sector of interest... ICT, Health, Engineering etc', default=None)
    parser.add_argument('-n', '--num', type=str, help='Number of jobs to be returned', required=False, default=10)
    parser.add_argument('-k', '--keyword', type=str, help='Keyword to search for', required=True)
    # Array for all arguments passed to script
    args = parser.parse_args()
    if args.state== None :
        args.state = 'Nigeria'
    # args = vars(arg)
    return args

def getSite(args):
  url = "http://www.careers24.com.ng/jobs/lc-{}".format(args.state)
  if args.keyword:
      url = url+"/kw-{}".format("-".join(args.keyword.split(',')))
  if args.sector:
      url = url+"/se-{}".format(args.sector)
  url = url+"/m-true/?sort=dateposted&pagesize={}".format(args.num)
  return get(url)

def getLinks(site):
  search_links = BeautifulSoup(site.text, "html.parser")
  return 'None' if len(search_links.findAll(text='No matches found'))!=0 else search_links.select("h4 > a")


if __name__ == "__main__":

    wb = load_workbook(filename = "Mass Job Upload.xlsx")
    active = wb["Fields"]
    start_row = 2
    col = 2
    args = get_args()
    log.debug(vars(args))
    log.debug(getLinks(getSite(args)))
    if getLinks(getSite(args)) == 'None':
        print 'No matches found for {} kindly search with another set of keywords'.format(args.keyword)

    site = getSite(args)
    if site.status_code == codes.ok:
        link_count = 0
        for link in getLinks(site):
          link_count += 1
          print "Link: ", link_count,
          url = 'http://www.careers24.com.ng'+link.attrs['href']
          page_url = get(url)
          page_content = BeautifulSoup(page_url.text, "html.parser")
          active.cell(column=col, row = 1, value=args.state)
          active.cell(column=col, row = 20, value="%s" % page_content.select('[itemprop="title"]')[0].getText())
          print page_content.select('[itemprop="title"]')[0].getText()
          active.cell(column=col, row = 7, value="%s" % page_content.select('[itemprop="baseSalary"]')[0].getText())

          description = page_content.select('[itemprop="hiringOrganization"] > ul')
          if len(description) > 0:
            active.cell(column=col, row = 21).value = ""
            for item in description:
              try:
                active.cell(column=col, row = 21).value += escape(item.getText())
              except Exception, e:
                pass

          requirements = page_content.select('[itemprop="responsibilities"] > ul')
          if len(requirements) > 0:
            active.cell(column=col, row = 23).value = ""
            for item in requirements:
              active.cell(column=col, row = 23).value += item.getText()
          active.cell(column=col, row = 24, value=url)
          col += 1
          time.sleep(1)
    else:
        print "Page Not Found"
    output_name = "%s-%d.xlsx" % (args.state.capitalize(), args.num)
    wb.save(output_name)

